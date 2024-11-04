import csv
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import ColorScaleRule

cols = {"method": 0, "accuracy":1, "AT":2, "time":3, "N_threads":4, "N_Epoch":5  }

name_methods = ["sequential", "multi_env", "synchrone", "synchrone_Imp",
                 "Asynchrone_1_2","Asynchrone_1_2_Imp", "Asynchrone_1_4", "Asynchrone_1_4_Imp"
                   "Adaptative_60", "Adaptative_60_Imp", "Adaptative_70", "Adaptative_70_Imp"
                    "Adaptative_80", "Adaptative_80_Imp", "N_synchrone_Imp", "N_synchrone_Imp_Mat",
                      "N_adaptative_60",   ]

best_methods = []

time_seq = 0
results = {}

def process_data(row):
    # Fonction simple : multiplier chaque élément par 2
    # Vous pouvez modifier cette fonction selon vos besoins
    return [float(x)  for x in row]

def get_min_time(data, id):
   
    if(id < len(data)):
        thread = int(data[id][cols["N_threads"]])
        if(float(data[id][cols["accuracy"]]) >= 0.95):
            time_min = int(data[id][cols["time"]])
        else:
            time_min = 1e40
        id_time_min = id        
        current_thread = thread
        
        while(current_thread == thread):
            if(time_min > int (data[id][cols["time"]])):
                if(float(data[id][cols["accuracy"]]) >= 0.95):
                    time_min = int (data[id][cols["time"]])
                    id_time_min = id
            id+=1
            if(id >= len(data)):    break
            current_thread = int(data[id][cols["N_threads"]])

    
        return id_time_min, id

    return None    


def process_csv(input_file, output_file):
    with open(input_file, 'r', newline='') as infile, open(output_file, 'w', newline='') as outfile:
        reader = list(csv.reader(infile))
        writer = csv.writer(outfile)
        print(type(list(reader)))
    
        id = 0
        current_method = 0
        time_seq = 0
        id_seq = 0
        for row in reader:
            if(id >= len(reader)):
                break
            id_write, i = get_min_time(reader, id)
            # print(type(row))
            # print(row)
            if(id_write!= None):
                if(float(reader[id][cols["accuracy"]]) >= 0.95):
                
                    processed_row = reader[id_write]
                    
                    if(id != 0):
                        processed_row.append(float(reader[id_seq][cols["time"]]) / float(reader[id_write][cols["time"]]))
                    else:
                        id_seq = id_write
                        processed_row.append(float(1))
                        
                    writer.writerow(processed_row)
                id = i
        



# Utilisation du script
input_file = 'speedup.csv'
output_file = 'output.csv'
process_csv(input_file, output_file)
print(f"Traitement terminé. Résultats écrits dans {output_file}")


def process_csv_to_excel(input_file, output_file):
    # Lire le CSV et traiter les données
    data = []
    with open(input_file, 'r', newline='') as infile:
        reader = csv.reader(infile)
        for row in reader:
            data.append(process_data(row))
        
    
    # Créer un nouveau classeur Excel et une feuille
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    # Écrire les données dans la feuille Excel
    for row in data:
        sheet.append(row)

    
    
    # Obtenir la plage de cellules contenant des données
    min_col = sheet.min_column
    max_col = sheet.max_column
    min_row = sheet.min_row
    max_row = sheet.max_row

    # Appliquer le formatage conditionnel à chaque colonne
    col_letter = openpyxl.utils.get_column_letter(7)
    sheet.conditional_formatting.add(
        f'{col_letter}1:{col_letter}{max_row}',
        ColorScaleRule(
            start_type='min', start_color='FFFFFF',
            mid_type='percentile', mid_value=50, mid_color='7FFF00',
            end_type='max', end_color='105d18'
        )
    )

    # Sauvegarder le fichier Excel
    workbook.save(output_file)

input_file = 'output.csv'
output_file = 'output.xlsx'
process_csv_to_excel(input_file, output_file)